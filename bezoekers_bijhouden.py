import tkinter as tk
from tkinter import messagebox
import pandas as pd
import csv
from datetime import datetime
import os
import locale
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

locale.setlocale(locale.LC_TIME, 'nl_NL.UTF-8')

def bezoekers_registratie():
    bezoekers = entry.get()
    
    if not bezoekers.isdigit():
        messagebox.showerror("Ongeldige invoer", "Voer een getal in.")
        return

    today = datetime.now()
    day_of_week = today.strftime('%A')
    file_path = 'bezoekers_registratie.xlsx'

    sheet_name = 'Bezoekers Registratie'

    new_data = pd.DataFrame({
            'Datum': [today.strftime('%d-%m-%Y')],
            'Dag': [day_of_week],
            'Aantal Bezoekers': [int(bezoekers)]
        })
   
     # Check if the Excel file already exists
    if os.path.exists(file_path):
        # Load the existing Excel file
        existing_data = pd.read_excel(file_path, engine='openpyxl')
        if today.strftime('%d-%m-%Y') in existing_data['Datum'].values:
                # Update existing entry
                updated_data = existing_data.loc[existing_data['Datum'] == today.strftime('%d-%m-%Y'), 'Bezoekers'] = [int(bezoekers)]
        else:
                # Append the new data
                updated_data = pd.concat([existing_data, new_data], ignore_index=True)
    else:
        # If the file doesn't exist, the new data is the full dataset
        updated_data = new_data

    # Save the updated data to the Excel file with the specified sheet name
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
        updated_data.to_excel(writer, sheet_name= sheet_name, index=False)

    
    # Apply formatting with openpyxl
    book = load_workbook(file_path)
    sheet = book[sheet_name]

    # Define styles
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(border_style='thin'), 
                         right=Side(border_style='thin'),
                         top=Side(border_style='thin'),
                         bottom=Side(border_style='thin'))

    # Apply header formatting
    for cell in sheet[1]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Apply formatting to all cells
    for row in sheet.iter_rows(min_row=2, max_col=len(updated_data.columns), max_row=sheet.max_row):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    column_widths = {'A': 15, 'B': 20, 'C': 30}  # Adjust widths as needed
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # Save the workbook
    book.save(file_path)

    messagebox.showinfo("Success", f"Er zijn {bezoekers} bezoekers geregistreerd op {day_of_week}.")
    entry.delete(0, tk.END)

# Create the main window
root = tk.Tk()
root.title("Bezoekers registratie")

# Configure grid to center the widgets
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=0)
root.grid_rowconfigure(2, weight=0)
root.grid_rowconfigure(3, weight=0)
root.grid_rowconfigure(4, weight=1)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)

# Create a Frame to hold the widgets
frame = tk.Frame(root)
frame.grid(row=1, column=0, columnspan=2, pady=20)

# Create and place the widgets
label = tk.Label(frame, text="Voer het aantal bezoekers in:")
label.grid(row=0, column=0, columnspan=2, padx=10, pady=5)

entry = tk.Entry(frame)
entry.grid(row=1, column=0, columnspan=2, padx=10, pady=5)

button = tk.Button(frame, text="Bezoekers registratie", command=bezoekers_registratie)
button.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

# Start the GUI event loop
root.mainloop()